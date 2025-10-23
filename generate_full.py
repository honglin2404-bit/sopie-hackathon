from openpyxl import load_workbook
import json

print("📖 Reading Excel with openpyxl...")

# Load workbook
wb = load_workbook('DB_ZaloPay_CS.xlsx', data_only=True)

# Generate Templates
ws_templates = wb['Response_Templates']
templates = []

# Get headers from first row
headers = []
for cell in ws_templates[1]:
    headers.append(cell.value)

print(f"Headers: {headers}")

# Read data rows
for row in ws_templates.iter_rows(min_row=2, values_only=True):
    template = {}
    for i, value in enumerate(row):
        if i < len(headers):
            # Convert None to null, keep everything else
            template[headers[i]] = value
    templates.append(template)

print(f"Read {len(templates)} templates")

# Save
with open('public/data/templates.json', 'w', encoding='utf-8') as f:
    json.dump(templates, f, ensure_ascii=False, indent=2)

print("✅ Saved to templates.json")